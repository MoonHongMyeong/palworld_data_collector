import openpyxl


def read_excel_to_list(file_name):
    workbook = openpyxl.load_workbook(file_name)
    ws = workbook.active

    result = []
    for rows in ws.iter_rows(min_col=1, max_col=ws.max_column):
        to_dict = {}
        offset_cnt = 0
        for cell in rows:
            if rows[0].value is not None:
                if cell.col_idx == 1:
                    breeding_base = cell.value
                    breeding_base_id = breeding_base.split('.')[0]
                    breeding_base_name = breeding_base.split('.')[1]
                    to_dict[breeding_base_id] = {}
                    to_dict[breeding_base_id]['no'] = breeding_base_id
                    to_dict[breeding_base_id]['name'] = breeding_base_name
                    to_dict[breeding_base_id]['mating'] = {}
                else:
                    breeding_base_id = cell.offset(row=0, column=-1 + offset_cnt).value.split('.')[0]
                    breeding_target_id = cell.offset(row=-1, column=0).value.split('.')[0]
                    to_dict[breeding_base_id]['mating'][breeding_target_id] = cell.value.split('.')[0]
                    offset_cnt -= 1
        if len(to_dict) != 0:
            result.append(to_dict)
    return result


if __name__ == '__main__':
    data = read_excel_to_list("data/data.xlsx")
    print(data)
