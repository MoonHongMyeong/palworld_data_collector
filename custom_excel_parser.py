import openpyxl


def read_excel(file_name):
    workbook = openpyxl.load_workbook(file_name)
    ws = workbook.active
    return ws


def to_dict_list(ws):
    result = []
    offset_row_cnt = 0
    for rows in ws.iter_rows(min_col=1, max_col=ws.max_column):
        to_dict = {}

        for cell in rows:
            if rows[0].value is not None:
                if cell.col_idx == 1:
                    breeding_base = cell.value
                    breeding_base_id = breeding_base.split('.')[0]
                    breeding_base_name = breeding_base.split('.')[1]
                    to_dict['no'] = breeding_base_id
                    to_dict['name'] = breeding_base_name.strip()
                    to_dict['mating'] = []
                else:
                    breeding_target = {}
                    breeding_target_id = cell.offset(row=-1 + offset_row_cnt, column=0).value.split('.')[0]
                    breeding_target['base'] = to_dict['no']
                    breeding_target['target'] = breeding_target_id
                    breeding_target['result'] = cell.value.split('.')[0]
                    to_dict['mating'].append(breeding_target)
                    # to_dict['mating']['target'].append({breeding_target_id: cell.value.split('.')[0]})

        if rows[0].value is not None:
            offset_row_cnt -= 1

        if len(to_dict) != 0:
            result.append(to_dict)
    return result

